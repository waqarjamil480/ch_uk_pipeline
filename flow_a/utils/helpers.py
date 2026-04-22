"""
utils/helpers.py
----------------
All non-DB utilities for Flow A consolidated into one file:
  - Config (dataclass)
  - Logging setup
  - Retry decorator
  - Downloader
  - CSV loader / row normaliser
  - URL resolver
  - Email notifier (Gmail)
  - Excel report writer
"""

import csv
import functools
import io
import logging
import os
import smtplib
import sys
import time
import traceback
import urllib.request
import zipfile
from dataclasses import dataclass, field
from datetime import date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Callable, Dict, Generator, List, Optional, Tuple

from dotenv import load_dotenv

# ── Load .env ─────────────────────────────────────────────────────────────
_ENV_FILE = Path(__file__).parent.parent / ".env"
load_dotenv(dotenv_path=_ENV_FILE, override=True)

logger = logging.getLogger(__name__)


# ===========================================================================
# CONFIG
# ===========================================================================

def _root() -> Path:
    return Path(__file__).parent.parent.resolve()


@dataclass
class Config:
    db_dsn: str = field(
        # DB_DSN is built from unified config.py (DB_HOST, DB_PORT, DB_NAME etc)
        # and passed explicitly from flow_a/main.py — default never used
        default_factory=lambda: ""
    )
    force_year: Optional[int] = field(
        default_factory=lambda: int(os.environ["FORCE_YEAR"]) if os.environ.get("FORCE_YEAR") else None
    )
    force_month: Optional[int] = field(
        default_factory=lambda: int(os.environ["FORCE_MONTH"]) if os.environ.get("FORCE_MONTH") else None
    )
    download_dir: str = field(
        default_factory=lambda: os.environ.get("DOWNLOAD_DIR", str(_root() / "downloads"))
    )
    report_dir: str = field(
        default_factory=lambda: os.environ.get("REPORT_DIR", str(_root() / "reports_uk_ch"))
    )
    batch_size: int = field(
        default_factory=lambda: int(os.environ.get("BATCH_SIZE", "500"))
    )
    parse_workers: int = field(
        default_factory=lambda: int(os.environ.get("PARSE_WORKERS", "8"))
    )
    db_workers: int = field(
        default_factory=lambda: int(os.environ.get("DB_WORKERS", "8"))
    )
    cleanup_downloads: bool = field(
        default_factory=lambda: os.environ.get("CLEANUP_DOWNLOADS", "true").lower() == "true"
    )
    max_retries: int = field(
        default_factory=lambda: int(os.environ.get("MAX_RETRIES", "3"))
    )
    retry_delay_seconds: int = field(
        default_factory=lambda: int(os.environ.get("RETRY_DELAY_SECONDS", "3"))
    )
    log_level: str = field(
        default_factory=lambda: os.environ.get("LOG_LEVEL", "INFO").upper()
    )


def load_config() -> Config:
    cfg = Config()
    Path(cfg.download_dir).mkdir(parents=True, exist_ok=True)
    Path(cfg.report_dir).mkdir(parents=True, exist_ok=True)
    return cfg


# ===========================================================================
# LOGGING
# ===========================================================================

def setup_logging(level: str = "INFO") -> None:
    numeric = getattr(logging, level.upper(), logging.INFO)
    fmt     = "%(asctime)s  %(levelname)-8s  %(name)s  %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"

    log_file = _root() / "pipeline.log"
    handlers = [
        RotatingFileHandler(str(log_file), maxBytes=10*1024*1024, backupCount=3, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ]
    for h in handlers:
        h.setFormatter(logging.Formatter(fmt=fmt, datefmt=datefmt))

    root = logging.getLogger()
    root.setLevel(numeric)
    root.handlers.clear()
    for h in handlers:
        root.addHandler(h)


# ===========================================================================
# RETRY DECORATOR
# ===========================================================================

def retry(max_attempts: int = 3, delay_seconds: float = 3.0) -> Callable:
    """Retry decorator — retries on any Exception up to max_attempts times."""
    def decorator(func: Callable) -> Callable:
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            last_err = None
            for attempt in range(1, max_attempts + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as exc:
                    last_err = exc
                    logger.warning(
                        "Pipeline attempt %d/%d failed: %s", attempt, max_attempts, exc
                    )
                    if attempt < max_attempts:
                        logger.info(
                            "Retrying in %.1f second(s) … (attempt %d/%d)",
                            delay_seconds, attempt + 1, max_attempts,
                        )
                        time.sleep(delay_seconds)
                    else:
                        logger.error("Pipeline exhausted all %d attempts — giving up.", max_attempts)
            raise last_err
        return wrapper
    return decorator


# ===========================================================================
# DOWNLOADER
# ===========================================================================

_CHUNK = 10 * 1024 * 1024  # 10 MB


def download_single(url: str, download_dir: str) -> Path:
    """Download URL to download_dir. Skips if already exists."""
    filename = url.split("/")[-1]
    dest     = Path(download_dir) / filename

    if dest.exists():
        logger.info("Already downloaded, skipping: %s", dest.name)
        return dest

    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(".tmp")

    logger.info("Downloading %s", url)
    try:
        with urllib.request.urlopen(url) as resp, open(tmp, "wb") as fh:
            total      = int(resp.headers.get("Content-Length", 0))
            downloaded = 0
            log_at     = _CHUNK
            while True:
                chunk = resp.read(_CHUNK)
                if not chunk:
                    break
                fh.write(chunk)
                downloaded += len(chunk)
                if downloaded >= log_at:
                    pct = (downloaded / total * 100) if total else 0
                    logger.info("  %.1f / %.1f MB  (%.0f%%)",
                                downloaded / 1e6, total / 1e6, pct)
                    log_at += _CHUNK
        tmp.rename(dest)
    except Exception:
        if tmp.exists():
            tmp.unlink()
        raise

    logger.info("Saved: %s  (%.1f MB)", dest, dest.stat().st_size / 1e6)
    return dest


def cleanup_zips(zip_paths: List[Path]) -> None:
    """Delete downloaded ZIPs after successful processing."""
    for path in zip_paths:
        try:
            if path.exists():
                path.unlink()
                logger.info("Deleted: %s", path)
        except Exception as exc:
            logger.warning("Could not delete %s: %s", path, exc)
    if zip_paths:
        folder = zip_paths[0].parent
        try:
            if folder.exists() and not any(folder.iterdir()):
                folder.rmdir()
                logger.info("Removed empty downloads folder: %s", folder)
        except Exception:
            pass


# ===========================================================================
# URL RESOLVER
# ===========================================================================

_BASE = "https://download.companieshouse.gov.uk"
_TMPL = "{base}/BasicCompanyDataAsOneFile-{year}-{month:02d}-01.zip"


def _url_exists(url: str, timeout: int = 10) -> bool:
    try:
        req = urllib.request.Request(url, method="HEAD")
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.status in (200, 206)
    except Exception:
        return False


def resolve_download_url(
    force_year: Optional[int] = None,
    force_month: Optional[int] = None,
) -> Tuple[str, int, int]:
    """Return (url, year, month) for the latest available CH snapshot."""
    if force_year and force_month:
        url = _TMPL.format(base=_BASE, year=force_year, month=force_month)
        if not _url_exists(url):
            raise RuntimeError(f"Forced URL not available: {url}")
        return url, force_year, force_month

    today = date.today()
    year, month = today.year, today.month
    for offset in range(3):
        m = month - offset
        y = year
        if m <= 0:
            m += 12
            y -= 1
        url = _TMPL.format(base=_BASE, year=y, month=m)
        logger.info("Checking: %s", url)
        if _url_exists(url):
            logger.info("Resolved → %s", url)
            return url, y, m
        logger.info("  Not available yet, checking prior month …")

    raise RuntimeError(
        f"No CH CSV found for {year}-{month:02d} or the 2 prior months."
    )


# ===========================================================================
# CSV LOADER
# ===========================================================================

CSV_COLUMN_MAP: Dict[str, str] = {
    " CompanyNumber":           "company_number",
    "CompanyName":              "company_name",
    "RegAddress.AddressLine1":  "address_line_1",
    "RegAddress.AddressLine2":  "address_line_2",
    "RegAddress.PostTown":      "post_town",
    "RegAddress.County":        "county",
    "RegAddress.Country":       "country",
    "RegAddress.PostCode":      "postcode",
    "RegAddress.CareOf":        "care_of",
    "RegAddress.POBox":         "po_box",
    "CompanyCategory":          "company_category",
    "CompanyStatus":            "company_status",
    "CountryOfOrigin":          "country_of_origin",
    "DissolutionDate":          "dissolution_date",
    "IncorporationDate":        "incorporation_date",
    "Accounts.AccountRefDay":   "accounts_ref_day",
    "Accounts.AccountRefMonth": "accounts_ref_month",
    "Accounts.NextDueDate":     "accounts_next_due_date",
    "Accounts.LastMadeUpDate":  "accounts_last_made_up",
    "Accounts.AccountCategory": "accounts_category",
    "Returns.NextDueDate":      "returns_next_due_date",
    "Returns.LastMadeUpDate":   "returns_last_made_up",
    "SICCode.SicText_1":        "sic_code_1",
    "SICCode.SicText_2":        "sic_code_2",
    "SICCode.SicText_3":        "sic_code_3",
    "SICCode.SicText_4":        "sic_code_4",
}

_ALL_COLS = {
    "company_number",
    "company_name",
    "address_line_1",
    "address_line_2",
    "post_town",
    "county",
    "country",
    "postcode",
    "care_of",
    "po_box",
    "company_category",
    "company_status",
    "country_of_origin",
    "dissolution_date",
    "incorporation_date",
    "accounts_ref_day",
    "accounts_ref_month",
    "accounts_next_due_date",
    "accounts_last_made_up",
    "accounts_category",
    "returns_next_due_date",
    "returns_last_made_up",
    "sic_code_1",
    "sic_code_2",
    "sic_code_3",
    "sic_code_4",
}


def _normalise_row(raw: Dict[str, str]) -> Dict[str, Any]:
    row: Dict[str, Any] = {col: None for col in _ALL_COLS}
    for csv_col, db_col in CSV_COLUMN_MAP.items():
        value = raw.get(csv_col)
        if value is not None:
            row[db_col] = value.strip() or None
    if row["company_number"]:
        row["company_number"] = str(row["company_number"]).strip()
    return row


def _detect_encoding(zip_path: Path, csv_name: str) -> str:
    with zipfile.ZipFile(zip_path, "r") as zf:
        with zf.open(csv_name) as rb:
            sample = rb.read(512 * 1024)
    try:
        sample.decode("utf-8-sig")
        return "utf-8-sig"
    except UnicodeDecodeError:
        return "cp1252"


def stream_rows_from_zip(
    zip_path: Path, batch_size: int = 500
) -> Generator[List[Dict[str, Any]], None, None]:
    """Stream rows from ZIP in batches for COPY into staging."""
    with zipfile.ZipFile(zip_path, "r") as zf:
        csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
        if not csv_names:
            raise ValueError(f"No CSV found inside {zip_path}")

    batch: List[Dict[str, Any]] = []
    row_count = 0

    for csv_name in csv_names:
        encoding = _detect_encoding(zip_path, csv_name)
        logger.info("Reading CSV: %s  (encoding=%s)", csv_name, encoding)
        with zipfile.ZipFile(zip_path, "r") as zf2:
            with zf2.open(csv_name) as rb:
                text   = io.TextIOWrapper(rb, encoding=encoding, errors="replace")
                reader = csv.DictReader(text)
                for raw_row in reader:
                    batch.append(_normalise_row(raw_row))
                    row_count += 1
                    if len(batch) >= batch_size:
                        yield batch
                        batch = []

    if batch:
        yield batch
    logger.info("Finished — %d rows", row_count)


# ===========================================================================
# EXCEL REPORT WRITER
# ===========================================================================

def write_report(
    report_dir: str,
    file_stats: List[Dict],
    db_table_counts: Dict[str, int],
    extra_stats: Optional[Dict] = None,
) -> Optional[Path]:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed — pip install openpyxl")
        return None

    H_FILL = PatternFill("solid", fgColor="1F4E79")
    H_FONT = Font(bold=True, color="FFFFFF")
    ALT    = PatternFill("solid", fgColor="D6E4F0")
    PASS   = PatternFill("solid", fgColor="C6EFCE")
    FAIL   = PatternFill("solid", fgColor="FFC7CE")

    def hdr(cell):
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center")

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = w + 4

    from datetime import datetime
    Path(report_dir).mkdir(parents=True, exist_ok=True)
    out = Path(report_dir) / f"company_data_report_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    wb  = openpyxl.Workbook()

    # Sheet 1: File Summary
    ws1 = wb.active; ws1.title = "File Summary"
    for col, h in enumerate(["File Name", "Rows Read", "Status", "Processed At"], 1):
        hdr(ws1.cell(1, col, h))
    for ri, s in enumerate(file_stats, 2):
        ws1.cell(ri, 1, s.get("file_name", ""))
        ws1.cell(ri, 2, s.get("row_count", 0)).number_format = "#,##0"
        ws1.cell(ri, 3, s.get("status", ""))
        ws1.cell(ri, 4, s.get("processed_at", ""))
        if ri % 2 == 0:
            for c in range(1, 5): ws1.cell(ri, c).fill = ALT
    tr = len(file_stats) + 2
    ws1.cell(tr, 1, "TOTAL").font = Font(bold=True)
    tc = ws1.cell(tr, 2, sum(s.get("row_count", 0) for s in file_stats))
    tc.number_format = "#,##0"; tc.font = Font(bold=True)
    autofit(ws1); ws1.freeze_panes = "A2"

    # Sheet 2: Load Stats
    es = extra_stats or {}
    ws2 = wb.create_sheet("Load Stats")
    for col, h in enumerate(["Metric", "Value", "Notes"], 1):
        hdr(ws2.cell(1, col, h))
    snap_ok = str(es.get("snapshot_check", "")).upper() == "OK"
    metrics = [
        ("Snapshot integrity check", es.get("snapshot_check", ""),
         "PASSED" if snap_ok else "FAILED — swap aborted"),
        ("Rows staged (COPY)", es.get("rows_staged", 0), "Bulk-loaded via COPY"),
        ("Rows upserted to metadata_uk_ch", es.get("rows_upserted", 0), "Live companies"),
    ]
    for ri, (label, val, note) in enumerate(metrics, 2):
        ws2.cell(ri, 1, label)
        cv = ws2.cell(ri, 2, val)
        ws2.cell(ri, 3, note)
        if isinstance(val, int): cv.number_format = "#,##0"
        if label == "Snapshot integrity check":
            fill = PASS if snap_ok else FAIL
            for c in range(1, 4): ws2.cell(ri, c).fill = fill
        elif ri % 2 == 0:
            for c in range(1, 4): ws2.cell(ri, c).fill = ALT
    autofit(ws2); ws2.freeze_panes = "A2"

    # Sheet 3: DB Table Counts
    ws3 = wb.create_sheet("DB Table Counts")
    for col, h in enumerate(["Table", "Row Count"], 1):
        hdr(ws3.cell(1, col, h))
    for ri, (table, count) in enumerate(db_table_counts.items(), 2):
        ws3.cell(ri, 1, table)
        ws3.cell(ri, 2, count).number_format = "#,##0"
        if ri % 2 == 0:
            for c in range(1, 3): ws3.cell(ri, c).fill = ALT
    autofit(ws3); ws3.freeze_panes = "A2"

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(out)
    logger.info("Excel report: %s", out)
    return out

# Email and Slack handled by shared utils/email_alert.py and utils/slack_alert.py
