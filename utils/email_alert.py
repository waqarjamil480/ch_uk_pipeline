"""
utils/email_alert.py
--------------------
Gmail email alerts + Slack notifications for Flow A & B pipeline.

Credentials are sourced from:
  - Airflow connection 'smtp_w1_default' (SMTP host/port/login/password)
  - 'receiver_email' stored in connection's Extra JSON field
    e.g.  {"receiver_email": "you@example.com"}

EMAIL_ENABLED / SLACK_ENABLED flags still come from .env.
"""

import csv
import io
import json
import logging
import os
import smtplib
import traceback
import urllib.request
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Dict, List, Optional

from dotenv import load_dotenv

_PROJECT_ROOT = Path(__file__).parent.parent.resolve()
load_dotenv(dotenv_path=_PROJECT_ROOT / ".env", override=False)

logger = logging.getLogger(__name__)


def _load_credentials():
    """Load SMTP credentials from Airflow connection, fallback to env vars."""
    try:
        from airflow.hooks.base import BaseHook
        conn = BaseHook.get_connection("smtp_w1_default")
        extra = {}
        if conn.extra:
            try:
                extra = json.loads(conn.extra)
            except Exception:
                pass
        return {
            "smtp_server":     conn.host or "smtp.gmail.com",
            "smtp_port":       int(conn.port) if conn.port else 587,
            "sender_email":    (conn.login or "").strip(),
            "sender_password": (conn.password or "").strip(),
            "receiver_email":  extra.get("receiver_email",
                                         os.getenv("RECEIVER_EMAIL", "")).strip(),
        }
    except Exception:
        return {
            "smtp_server":     os.getenv("SMTP_SERVER",     "smtp.gmail.com"),
            "smtp_port":       int(os.getenv("SMTP_PORT",   "587")),
            "sender_email":    os.getenv("SENDER_EMAIL",    "").strip(),
            "sender_password": os.getenv("SENDER_PASSWORD", "").strip(),
            "receiver_email":  os.getenv("RECEIVER_EMAIL",  "").strip(),
        }


# Flags from .env
EMAIL_ENABLED     = os.getenv("EMAIL_ENABLED",   "false").lower() == "true"
SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL", "").strip()
SLACK_ENABLED     = os.getenv("SLACK_ENABLED",     "false").lower() == "true"


# =============================================================================
# PUBLIC API
# =============================================================================

def notify(
    status: str,
    month: str,
    processed: int = 0,
    failed: int = 0,
    duration=None,
    error: Exception = None,
    failed_files:  Optional[List[Dict]] = None,
    warning_files: Optional[List[Dict]] = None,
    flow: str = "Flow B",
    extra_stats: Optional[Dict] = None,
) -> None:
    _send_email_alert(
        status=status, month=month, processed=processed,
        failed=failed, duration=duration, error=error,
        failed_files=failed_files, warning_files=warning_files,
        flow=flow, extra_stats=extra_stats,
    )
    _send_slack_alert(
        status=status, month=month, processed=processed,
        failed=failed, duration=duration, error=error,
        failed_files=failed_files, warning_files=warning_files,
        flow=flow,
    )


# =============================================================================
# EMAIL
# =============================================================================

def _send_email_alert(
    status, month, processed, failed, duration, error,
    failed_files, warning_files, flow="Flow B",
    extra_stats=None,
):
    if not EMAIL_ENABLED:
        logger.info("Email disabled (EMAIL_ENABLED=false)")
        return

    creds = _load_credentials()
    if not creds["sender_email"] or not creds["sender_password"] or not creds["receiver_email"]:
        logger.warning("Email skipped — SMTP credentials not set in Airflow connection 'smtp_w1_default'")
        return

    if status == "success":
        subject = f"✅ CH Pipeline SUCCESS — {flow} | {month}"
        body    = _email_success_html(month, processed, failed, duration,
                                      flow=flow, extra_stats=extra_stats)
    elif status == "failure":
        subject = f"❌ CH Pipeline FAILURE — {flow} | {month}"
        body    = _email_failure_html(month, processed, failed, duration,
                                      error, flow=flow)
    else:
        return

    attachment      = None
    attachment_name = None
    if failed_files or warning_files:
        attachment      = _build_excel(failed_files or [], warning_files or [], month)
        attachment_name = f"pipeline_report_{month.replace('-', '_')}.xlsx"
        logger.info("Attaching Excel: %s", attachment_name)

    _smtp_send(creds, subject, body, attachment, attachment_name)


_CSS = """
body{font-family:Arial,sans-serif;color:#333;margin:0;padding:0}
.hdr{padding:20px 30px;color:#fff}
.body{padding:24px 30px}
table{border-collapse:collapse;width:100%;margin-top:16px}
th{background:#1F4E79;color:#fff;padding:10px 14px;text-align:left}
td{padding:9px 14px;border-bottom:1px solid #e0e0e0}
tr:nth-child(even) td{background:#f4f8fc}
.mono{font-family:monospace;font-size:13px;background:#f5f5f5;
      padding:14px;border-left:4px solid #ccc;white-space:pre-wrap;word-break:break-all}
.ftr{font-size:12px;color:#888;padding:12px 30px;border-top:1px solid #eee}
"""

def _r(label, value):
    return f"<tr><td>{label}</td><td>{value}</td></tr>"

def _email_success_html(month, processed, failed, duration,
                        flow="Flow B", extra_stats=None):
    if flow == "Flow A":
        es = extra_stats or {}
        stats_rows = f"""
  {_r("Rows staged via COPY",     f"{es.get('rows_staged', 0):,}")}
  {_r("Rows upserted to metadata_uk_ch",f"{es.get('rows_upserted', 0):,}")}
  {_r("Snapshot integrity check", es.get('snapshot_check', 'OK'))}
  {_r("Duration",                 str(duration) if duration else "N/A")}"""
    else:
        total = processed + failed
        rate  = f"{processed/total*100:.2f}%" if total > 0 else "N/A"
        stats_rows = f"""
  {_r("Files processed", f"{processed:,}")}
  {_r("Files failed",    f"{failed:,}")}
  {_r("Success rate",    rate)}
  {_r("Duration",        str(duration) if duration else "N/A")}"""

    return f"""<html><head><style>{_CSS}</style></head><body>
<div class="hdr" style="background:#1a7a3c;">
  <h2 style="margin:0">✅ Pipeline Completed Successfully</h2>
  <p style="margin:4px 0 0">Flow: <strong>{flow}</strong> &nbsp;|&nbsp; Month: <strong>{month}</strong></p>
</div>
<div class="body"><table>
  <tr><th colspan="2">Run Summary</th></tr>
  {_r("Flow",   f"<strong>{flow}</strong>")}
  {_r("Month",  f"<strong>{month}</strong>")}
  {_r("Status", "<span style='color:#1a7a3c;font-weight:bold'>SUCCESS</span>")}
  {stats_rows}
</table></div>
<div class="ftr">Companies House Data Pipeline — automated alert</div>
</body></html>"""

def _email_failure_html(month, processed, failed, duration,
                        error, flow="Flow B"):
    err_type = type(error).__name__ if error else "Unknown"
    err_msg  = str(error)          if error else "No details"
    tb       = traceback.format_exc()
    return f"""<html><head><style>{_CSS}</style></head><body>
<div class="hdr" style="background:#c0392b;">
  <h2 style="margin:0">❌ Pipeline Failed</h2>
  <p style="margin:4px 0 0">Flow: <strong>{flow}</strong> &nbsp;|&nbsp; Month: <strong>{month}</strong></p>
</div>
<div class="body"><table>
  <tr><th colspan="2">Run Summary</th></tr>
  {_r("Flow",                f"<strong>{flow}</strong>")}
  {_r("Month",               f"<strong>{month}</strong>")}
  {_r("Status",              "<span style='color:#c0392b;font-weight:bold'>FAILED</span>")}
  {_r("Duration",            str(duration) if duration else "N/A")}
  <tr><th colspan="2">Error Details</th></tr>
  {_r("Error type",    f"<strong>{err_type}</strong>")}
  {_r("Error message", err_msg)}
</table>
<p style="margin-top:20px"><strong>Full traceback:</strong></p>
<div class="mono">{tb}</div>
</div>
<div class="ftr">Companies House Data Pipeline — automated alert</div>
</body></html>"""


def _build_excel(failed_files, warning_files, month):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed — pip install openpyxl")
        return None

    def hdr(ws, row, col, value, bg):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center")

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 80)

    def fill_sheet(ws, rows, bg):
        hdr(ws, 1, 1, "ch_upload",        bg)
        hdr(ws, 1, 2, "source_file",      bg)
        hdr(ws, 1, 3, "error / warning",  bg)
        alt = PatternFill("solid", fgColor="F4F8FC")
        for ri, row in enumerate(rows, 2):
            ws.cell(ri, 1, row.get("ch_upload",   month))
            ws.cell(ri, 2, row.get("source_file", ""))
            ws.cell(ri, 3, row.get("comment",     ""))
            if ri % 2 == 0:
                for c in range(1, 4): ws.cell(ri, c).fill = alt
        tr = len(rows) + 2
        ws.cell(tr, 1, "TOTAL").font = Font(bold=True)
        ws.cell(tr, 2, len(rows)).font = Font(bold=True)
        autofit(ws)
        ws.freeze_panes = "A2"

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Failed Files"
    ws1.sheet_tab_color = "C0392B"
    fill_sheet(ws1, failed_files, "C0392B")

    ws2 = wb.create_sheet("Warning Files")
    ws2.sheet_tab_color = "E67E22"
    fill_sheet(ws2, warning_files, "E67E22")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _smtp_send(creds, subject, html_body, attachment=None, attachment_name=None):
    recipients = [r.strip() for r in creds["receiver_email"].split(",") if r.strip()]
    msg = MIMEMultipart("mixed" if attachment else "alternative")
    msg["Subject"] = subject
    msg["From"]    = f"CH Pipeline <{creds['sender_email']}>"
    msg["To"]      = ", ".join(recipients)
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    if attachment and attachment_name:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_name}"')
        msg.attach(part)

    try:
        with smtplib.SMTP(creds["smtp_server"], creds["smtp_port"], timeout=30) as smtp:
            smtp.ehlo(); smtp.starttls()
            smtp.login(creds["sender_email"], creds["sender_password"])
            smtp.sendmail(creds["sender_email"], recipients, msg.as_string())
        logger.info("Email sent → %s  [%s]", creds["receiver_email"], subject)
        if attachment_name:
            logger.info("  Attachment: %s", attachment_name)
    except smtplib.SMTPAuthenticationError:
        logger.error("SMTP auth failed — check credentials in Airflow connection 'smtp_w1_default'")
    except Exception as exc:
        logger.error("Email failed: %s", exc)


# =============================================================================
# SLACK
# =============================================================================

def _send_slack_alert(
    status, month, processed, failed, duration, error,
    failed_files, warning_files, flow="Flow B",
):
    if not SLACK_ENABLED:
        return
    if not SLACK_WEBHOOK_URL:
        logger.warning("Slack skipped — SLACK_WEBHOOK_URL not set in .env")
        return

    total    = processed + failed
    rate     = f"{processed/total*100:.2f}%" if total > 0 else "N/A"
    duration_str = str(duration) if duration else "N/A"
    n_failed  = len(failed_files)  if failed_files  else 0
    n_warning = len(warning_files) if warning_files else 0

    if status == "success":
        header_emoji = "✅"
        header_text  = f"*CH Pipeline Completed — {month}*"
        color        = "#1a7a3c"
        status_text  = "✅ SUCCESS"
        fields = [
            {"title": "Month",            "value": month,           "short": True},
            {"title": "Status",           "value": status_text,     "short": True},
            {"title": "Files Processed",  "value": f"{processed:,}","short": True},
            {"title": "Files Failed",     "value": f"{failed:,}",   "short": True},
            {"title": "Files Warning",    "value": f"{n_warning:,}","short": True},
            {"title": "Success Rate",     "value": rate,            "short": True},
            {"title": "Duration",         "value": duration_str,    "short": True},
        ]
        extra_text = ""
        if n_failed > 0:
            extra_text += f"\n⚠️ *{n_failed} files failed* — check Excel attachment in email"
        if n_warning > 0:
            extra_text += f"\n🔔 *{n_warning} files had warnings* — data saved, minor fields skipped"
        fallback = f"CH Pipeline SUCCESS — {month} | {processed:,} processed | {failed:,} failed | {duration_str}"
    else:
        header_emoji = "❌"
        header_text  = f"*CH Pipeline FAILED — {month}*"
        color        = "#c0392b"
        status_text  = "❌ FAILED"
        err_type = type(error).__name__ if error else "Unknown"
        err_msg  = str(error)[:200]     if error else "No details"
        tb_lines = traceback.format_exc().strip().split("\n")
        non_hl   = [l for l in tb_lines if l.strip() and not all(c in "~^ " for c in l.strip())]
        tb_short = "\n".join(non_hl[-3:]) if len(non_hl) >= 3 else "\n".join(non_hl)
        fields = [
            {"title": "Month",             "value": month,          "short": True},
            {"title": "Status",            "value": status_text,    "short": True},
            {"title": "Files Processed",   "value": f"{processed:,}","short": True},
            {"title": "Files Failed",      "value": f"{failed:,}",  "short": True},
            {"title": "Duration",          "value": duration_str,   "short": True},
            {"title": "Error Type",        "value": err_type,       "short": True},
            {"title": "Error Message",     "value": err_msg,        "short": False},
        ]
        extra_text = f"\n```{tb_short}```"
        fallback   = f"CH Pipeline FAILED — {month} | {err_type}: {err_msg}"

    payload = {
        "text": f"{header_emoji} {header_text}",
        "attachments": [
            {
                "fallback":    fallback,
                "color":       color,
                "fields":      fields,
                "footer":      "Companies House Data Pipeline",
                "footer_icon": "https://companieshouse.gov.uk/favicon.ico",
                "ts":          int(__import__("time").time()),
            }
        ],
    }
    if extra_text:
        payload["attachments"].append({"color": color, "text": extra_text})

    try:
        data = json.dumps(payload).encode("utf-8")
        req  = urllib.request.Request(
            SLACK_WEBHOOK_URL,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = resp.read().decode()
            if result == "ok":
                logger.info("Slack notification sent [%s]", month)
            else:
                logger.warning("Slack unexpected response: %s", result)
    except Exception as exc:
        logger.warning("Slack notification failed: %s", exc)
